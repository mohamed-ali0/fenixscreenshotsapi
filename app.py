from flask import Flask, jsonify, request, send_file, send_from_directory
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
import os
import json
import zipfile
from pathlib import Path
from automation import download_excel_report
from system_settings import SystemSettings
import threading
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import subprocess

app = Flask(__name__)

# Initialize system settings
settings = SystemSettings()

def add_separator_lines_to_excel(excel_path, output_path):
    """
    Add visible separator lines to Excel file and convert to PDF
    """
    try:
        print(f"[{datetime.now()}] Adding separator lines to Excel file...")
        
        # Load the Excel file
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        # Define border style for separator lines
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        # Save the modified Excel file
        workbook.save(output_path)
        print(f"[{datetime.now()}] Excel file with separator lines saved: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"[{datetime.now()}] Error adding separator lines: {str(e)}")
        return False

def convert_excel_to_pdf(excel_path, pdf_path):
    """
    Convert Excel file to PDF with borders in landscape legal format using PowerShell
    """
    try:
        print(f"[{datetime.now()}] Converting Excel to PDF with borders...")
        
        # Create PowerShell script for Excel to PDF conversion with borders
        ps_script = f'''
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false
        
        try {{
            $workbook = $excel.Workbooks.Open('{os.path.abspath(excel_path)}')
            $worksheet = $workbook.Worksheets.Item(1)
            
            # Add borders to all cells with data
            $usedRange = $worksheet.UsedRange
            if ($usedRange) {{
                $usedRange.Borders.LineStyle = 1  # xlContinuous
                $usedRange.Borders.Weight = 1     # xlThin
                # Set gray color for all border sides
                $usedRange.Borders.Item(7).Color = 12632256  # xlEdgeLeft - Light gray
                $usedRange.Borders.Item(8).Color = 12632256  # xlEdgeTop - Light gray  
                $usedRange.Borders.Item(9).Color = 12632256  # xlEdgeBottom - Light gray
                $usedRange.Borders.Item(10).Color = 12632256 # xlEdgeRight - Light gray
            }}
            
            # Set to Legal landscape
            $worksheet.PageSetup.Orientation = 2  # xlLandscape
            $worksheet.PageSetup.PaperSize = 5     # xlPaperLegal
            
            # Export to PDF
            $pdf_path = '{os.path.abspath(pdf_path)}'
            $workbook.ExportAsFixedFormat(0, $pdf_path)
            
            $workbook.Close()
            $excel.Quit()
            
            Write-Host "PDF created: $pdf_path"
        }} catch {{
            Write-Host "Error: $_"
            $excel.Quit()
        }}
        '''
        
        # Write PowerShell script to file
        with open('temp.ps1', 'w') as f:
            f.write(ps_script)
        
        try:
            # Run PowerShell script
            result = subprocess.run(['powershell', '-ExecutionPolicy', 'Bypass', '-File', 'temp.ps1'], 
                                  capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0:
                print(f"[{datetime.now()}] PDF conversion successful: {pdf_path}")
                return True
            else:
                print(f"[{datetime.now()}] PowerShell error: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            print(f"[{datetime.now()}] PDF conversion timed out")
            return False
        except Exception as e:
            print(f"[{datetime.now()}] Error: {e}")
            return False
        finally:
            # Clean up
            if os.path.exists('temp.ps1'):
                os.remove('temp.ps1')
            
    except Exception as e:
        print(f"[{datetime.now()}] Error in PDF conversion: {str(e)}")
        return False

# Initialize screenshots directory
SCREENSHOTS_DIR = "screenshots"
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

# Scheduler for automated screenshots
scheduler = BackgroundScheduler()
screenshot_lock = threading.Lock()


def scheduled_excel_download_task():
    """Task to capture Fenix Marine Services screenshots - runs according to frequency"""
    with screenshot_lock:
        try:
            print(f"[{datetime.now()}] Starting scheduled Fenix Marine Services screenshot capture...")
            credentials = settings.get_login_credentials()
            success, message = download_excel_report(
                credentials['username'], 
                credentials['password']
            )
            if success:
                print(f"[{datetime.now()}] Screenshot capture completed successfully: {message}")
            else:
                print(f"[{datetime.now()}] Screenshot capture failed: {message}")
                # Schedule retry in 5 minutes
                print(f"[{datetime.now()}] Scheduling retry in 5 minutes...")
                scheduler.add_job(
                    func=scheduled_excel_download_task,
                    trigger="date",
                    run_date=datetime.now() + timedelta(minutes=5),
                    id='screenshot_retry',
                    replace_existing=True
                )
        except Exception as e:
            print(f"[{datetime.now()}] Error in scheduled screenshot: {str(e)}")
            # Schedule retry in 5 minutes on exception
            print(f"[{datetime.now()}] Scheduling retry in 5 minutes after exception...")
            try:
                scheduler.add_job(
                    func=scheduled_excel_download_task,
                    trigger="date",
                    run_date=datetime.now() + timedelta(minutes=5),
                    id='screenshot_retry',
                    replace_existing=True
                )
            except:
                pass


def restart_scheduler():
    """Restart the scheduler with updated frequency and preferred hour"""
    scheduler.remove_all_jobs()
    frequency_hours = settings.get_frequency()
    preferred_hour = settings.get_preferred_hour()
    
    # Calculate next run time based on preferred hour
    now = datetime.now()
    next_run = now.replace(hour=preferred_hour, minute=0, second=0, microsecond=0)
    
    # If preferred hour has already passed today, schedule for next occurrence
    if next_run <= now:
        if frequency_hours >= 24:
            # Daily - schedule for tomorrow at preferred hour
            next_run = next_run + timedelta(days=1)
        else:
            # Multiple times per day - schedule for next occurrence
            # Find next valid time that's a multiple of frequency_hours from preferred hour
            hours_until_next = frequency_hours - ((now.hour - preferred_hour) % frequency_hours)
            if hours_until_next == frequency_hours:
                hours_until_next = 0
            next_run = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=hours_until_next)
            # If we went past midnight, set to preferred hour next day
            if next_run.hour != preferred_hour and next_run.day != now.day:
                next_run = next_run.replace(hour=preferred_hour)
    
    # Use cron trigger for daily, interval for multiple times per day
    if frequency_hours >= 24:
        # Daily at preferred hour
        scheduler.add_job(
            func=scheduled_excel_download_task,
            trigger="cron",
            hour=preferred_hour,
            minute=0,
            id='screenshot_job',
            replace_existing=True
        )
        print(f"Scheduler restarted: Daily at {preferred_hour:02d}:00")
    else:
        # Multiple times per day - use interval starting at preferred hour
        scheduler.add_job(
            func=scheduled_excel_download_task,
            trigger="interval",
            hours=frequency_hours,
            start_date=next_run,
            id='screenshot_job',
            replace_existing=True
        )
        print(f"Scheduler restarted: Every {frequency_hours} hours starting at {next_run.strftime('%H:%M')} (preferred: {preferred_hour:02d}:00)")


@app.route('/')
def index():
    """API information endpoint"""
    return jsonify({
        "name": "Fenix Marine Services Screenshot API",
        "version": "2.0.0",
        "endpoints": {
            "GET /screenshot/<date>": "Get screenshot by date (format: YYYY-MM-DD)",
            "GET /screenshots/range": "Get screenshots range (params: start_date, end_date or last_n)", 
            "GET /download/<filename>": "Download specific screenshot file",
            "POST /screenshot/now": "Capture full desktop screenshot immediately (requires admin_password)",
            "POST /admin/frequency": "Change screenshot frequency (requires admin_password, frequency_hours)",
            "POST /admin/preferred_hour": "Set preferred capture hour (requires admin_password, preferred_hour)",
            "POST /admin/credentials": "Update Fenix login credentials (requires admin_password, username, password)",
            "POST /admin/cleanup": "Delete all screenshots (requires admin_password)",
            "GET /status": "Get system status and statistics"
        }
    })


@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """
    Download a specific screenshot file
    """
    try:
        # Check screenshots directory first (primary location)
        if os.path.exists(SCREENSHOTS_DIR):
            file_path = os.path.join(SCREENSHOTS_DIR, filename)
            if os.path.exists(file_path):
                return send_from_directory(
                    SCREENSHOTS_DIR, 
                    filename,
                    as_attachment=True,
                    download_name=filename
                )
        
        # Check downloads directory as fallback (for any legacy files)
        downloads_dir = "downloads"
        if os.path.exists(downloads_dir):
            file_path = os.path.join(downloads_dir, filename)
            if os.path.exists(file_path):
                return send_from_directory(
                    downloads_dir, 
                    filename,
                    as_attachment=True,
                    download_name=filename
                )
        
        return jsonify({
            "success": False,
            "error": "Screenshot file not found"
        }), 404
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/screenshot/<date_str>', methods=['GET'])
def get_screenshot(date_str):
    """
    Get screenshot for a specific date
    Date format: YYYY-MM-DD
    Returns JSON with download link
    """
    try:
        # Validate date format
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        
        # Find screenshot files for that date
        date_pattern = target_date.strftime("%Y-%m-%d")
        matching_files = []
        
        # Check screenshots directory
        if os.path.exists(SCREENSHOTS_DIR):
            for filename in os.listdir(SCREENSHOTS_DIR):
                if filename.startswith(f"fenix_screenshot_{date_pattern}") and filename.endswith('.png'):
                    matching_files.append(filename)
        
        if not matching_files:
            return jsonify({
                "success": False,
                "error": f"No screenshot found for date {date_str}"
            }), 404
        
        # Return the most recent screenshot file for that date
        matching_files.sort(reverse=True)
        screenshot_file = matching_files[0]
        
        # Return JSON with download link
        download_url = f"{request.host_url}download/{screenshot_file}"
        
        return jsonify({
            "success": True,
            "date": date_str,
            "filename": screenshot_file,
            "download_url": download_url,
            "message": f"Screenshot found for {date_str}"
        })
    
    except ValueError:
        return jsonify({
            "success": False,
            "error": "Invalid date format. Use YYYY-MM-DD"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500




@app.route('/screenshots/range', methods=['GET'])
def get_screenshots_range():
    """
    Get screenshots for a date range or last N screenshots
    Parameters:
    - start_date: YYYY-MM-DD (required if not using last_n)
    - end_date: YYYY-MM-DD (required if not using last_n)
    - last_n: number of most recent screenshots (optional)
    Returns JSON with download link for ZIP file
    """
    try:
        last_n = request.args.get('last_n', type=int)
        
        if last_n:
            # Get last N screenshots
            all_files = sorted(
                [f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')],
                reverse=True
            )
            selected_files = all_files[:last_n]
        else:
            # Get screenshots in date range
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')
            
            if not start_date_str or not end_date_str:
                return jsonify({
                    "success": False,
                    "error": "Provide either 'last_n' or both 'start_date' and 'end_date'"
                }), 400
            
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            
            if start_date > end_date:
                return jsonify({
                    "success": False,
                    "error": "start_date must be before or equal to end_date"
                }), 400
            
            # Find all screenshots in range
            selected_files = []
            for filename in os.listdir(SCREENSHOTS_DIR):
                if filename.endswith('.png'):
                    try:
                        # Extract date from filename (format: YYYY-MM-DD_HH-MM-SS.png)
                        file_date_str = filename.split('_')[0]
                        file_date = datetime.strptime(file_date_str, "%Y-%m-%d")
                        
                        if start_date <= file_date <= end_date:
                            selected_files.append(filename)
                    except:
                        continue
            
            selected_files.sort()
        
        if not selected_files:
            return jsonify({
                "success": False,
                "error": "No screenshots found for the specified criteria"
            }), 404
        
        # Create zip file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"screenshots_{timestamp}.zip"
        zip_path = os.path.join(SCREENSHOTS_DIR, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for filename in selected_files:
                file_path = os.path.join(SCREENSHOTS_DIR, filename)
                zipf.write(file_path, filename)
        
        # Return JSON with download link
        download_url = f"{request.host_url}download/{zip_filename}"
        
        return jsonify({
            "success": True,
            "zip_filename": zip_filename,
            "download_url": download_url,
            "screenshot_count": len(selected_files),
            "screenshots": selected_files,
            "message": f"ZIP file created with {len(selected_files)} screenshot(s)"
        })
    
    except ValueError as e:
        return jsonify({
            "success": False,
            "error": "Invalid date format. Use YYYY-MM-DD"
        }), 400
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/frequency', methods=['POST'])
def change_frequency():
    """
    Change screenshot frequency
    Body: {
        "admin_password": "password",
        "frequency_hours": 24
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        frequency_hours = data.get('frequency_hours')
        
        if not admin_password or frequency_hours is None:
            return jsonify({
                "success": False,
                "error": "admin_password and frequency_hours are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Validate frequency
        if not isinstance(frequency_hours, (int, float)) or frequency_hours <= 0:
            return jsonify({
                "success": False,
                "error": "frequency_hours must be a positive number"
            }), 400
        
        # Update frequency
        settings.set_frequency(frequency_hours)
        restart_scheduler()
        
        return jsonify({
            "success": True,
            "message": f"Frequency updated to {frequency_hours} hours",
            "frequency_hours": frequency_hours
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/credentials', methods=['POST'])
def update_credentials():
    """
    Update login credentials for the portal
    Body: {
        "admin_password": "password",
        "username": "new_username",
        "password": "new_password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        username = data.get('username')
        password = data.get('password')
        
        if not admin_password or not username or not password:
            return jsonify({
                "success": False,
                "error": "admin_password, username, and password are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Update credentials
        settings.set_login_credentials(username, password)
        
        return jsonify({
            "success": True,
            "message": "Login credentials updated successfully"
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/cleanup', methods=['POST'])
def cleanup():
    """
    Delete all screenshots
    Body: {
        "admin_password": "password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        
        if not admin_password:
            return jsonify({
                "success": False,
                "error": "admin_password is required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Delete all screenshots
        deleted_count = 0
        for filename in os.listdir(SCREENSHOTS_DIR):
            if filename.endswith('.png') or filename.endswith('.zip'):
                file_path = os.path.join(SCREENSHOTS_DIR, filename)
                try:
                    os.remove(file_path)
                    deleted_count += 1
                except Exception as e:
                    print(f"Error deleting {filename}: {str(e)}")
        
        return jsonify({
            "success": True,
            "message": f"Cleanup completed. {deleted_count} files deleted."
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/admin/preferred_hour', methods=['POST'])
def set_preferred_hour():
    """
    Set preferred hour for scheduled captures
    Body: {
        "admin_password": "password",
        "preferred_hour": 10
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        preferred_hour = data.get('preferred_hour')
        
        if not admin_password or preferred_hour is None:
            return jsonify({
                "success": False,
                "error": "admin_password and preferred_hour are required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Validate hour
        if not isinstance(preferred_hour, int) or preferred_hour < 0 or preferred_hour > 23:
            return jsonify({
                "success": False,
                "error": "preferred_hour must be an integer between 0 and 23"
            }), 400
        
        # Update preferred hour
        settings.set_preferred_hour(preferred_hour)
        
        # Restart scheduler with new preferred hour
        restart_scheduler()
        
        return jsonify({
            "success": True,
            "message": f"Preferred hour updated to {preferred_hour}:00",
            "preferred_hour": preferred_hour
        })
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/screenshot/now', methods=['POST'])
def capture_screenshot_now():
    """
    Capture Fenix Marine Services screenshot immediately
    Body: {
        "admin_password": "password"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Request body must be JSON"
            }), 400
        
        admin_password = data.get('admin_password')
        
        if not admin_password:
            return jsonify({
                "success": False,
                "error": "admin_password is required"
            }), 400
        
        # Verify admin password
        if not settings.verify_admin_password(admin_password):
            return jsonify({
                "success": False,
                "error": "Invalid admin password"
            }), 403
        
        # Capture screenshot
        with screenshot_lock:
            credentials = settings.get_login_credentials()
            success, message = download_excel_report(
                credentials['username'], 
                credentials['password']
            )
        
        if success:
            # Extract filename from message if it contains a path
            if "screenshots" in message:
                filename = os.path.basename(message)
                return jsonify({
                    "success": True,
                    "message": "Fenix Marine Services screenshot captured successfully",
                    "filename": filename,
                    "download_url": f"{request.host_url}download/{filename}"
                })
            else:
                return jsonify({
                    "success": True,
                    "message": message
                })
        else:
            return jsonify({
                "success": False,
                "error": message
            }), 500
    
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/status', methods=['GET'])
def status():
    """Get current system status"""
    try:
        frequency = settings.get_frequency()
        preferred_hour = settings.get_preferred_hour()
        credentials = settings.get_login_credentials()
        
        # Count screenshot files
        screenshot_count = len([f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')])
        
        # Get last screenshot info
        screenshots = sorted(
            [f for f in os.listdir(SCREENSHOTS_DIR) if f.endswith('.png')],
            reverse=True
        )
        last_screenshot = screenshots[0] if screenshots else None
        
        # Get last screenshot date if available
        last_screenshot_date = None
        if last_screenshot:
            try:
                # Extract date from filename: fenix_screenshot_YYYY-MM-DD_HH-MM-SS.png
                date_part = last_screenshot.replace('fenix_screenshot_', '').split('_')[0]
                last_screenshot_date = date_part
            except:
                pass
        
        return jsonify({
            "success": True,
            "system_info": {
                "name": "Fenix Marine Services Screenshot API",
                "version": "2.0.0",
                "status": "running"
            },
            "scheduler_info": {
                "running": scheduler.running,
                "frequency_hours": frequency,
                "preferred_hour": preferred_hour
            },
            "credentials": {
                "username": credentials['username'],
                "portal": "portal.fenixmarineservices.com"
            },
            "screenshot_stats": {
                "total_screenshots": screenshot_count,
                "last_screenshot_file": last_screenshot,
                "last_screenshot_date": last_screenshot_date,
                "screenshots_directory": SCREENSHOTS_DIR
            }
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


if __name__ == '__main__':
    # Start the scheduler
    scheduler.start()
    restart_scheduler()
    
    print("=" * 50)
    print("Fenix Marine Services Screenshot API")
    print("=" * 50)
    print(f"Screenshot frequency: {settings.get_frequency()} hours")
    print(f"Preferred start time: {settings.get_preferred_hour():02d}:00")
    print(f"Screenshots directory: {SCREENSHOTS_DIR}")
    print("Scheduler started")
    print("=" * 50)
    
    try:
        app.run(host='0.0.0.0', port=5005, debug=False)
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()
        print("\nShutdown complete")

